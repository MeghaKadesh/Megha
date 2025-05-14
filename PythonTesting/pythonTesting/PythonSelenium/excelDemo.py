import openpyxl                  # download in commandpormt first import oepnpyxl
book = openpyxl.load_workbook("C:\\Users\\kades\\OneDrive\\Documents\\pythonDemo.xlsx")      # load excel and give path of that excel
sheet = book.active
Dist ={}      # empty distornary
# to open actve sheet
cell = sheet.cell(row=1, column=2)         # choose the particular cell
print(cell.value)                          # printing the value of that cell

#or
print(sheet['A3'].value)

sheet.cell(row=2,column=2).value ="Megha"   # choose one cell and add value inside that cell
print(sheet.cell(row=2,column=2).value)

print(sheet.max_row)                 # maximun row of the excel
print(sheet.max_column)              # maimun  column of the excel

for i in range(1,sheet.max_row+1):    # printing all data  from the excel
    if sheet.cell(row=i,column=1).value == "Testcase2":
        for j in range(1,sheet.max_column+1):
            Dist[sheet.cell(row=1,column=j).value]=sheet.cell(row=i, column=j).value   # printing inside the distonary
print(Dist)



